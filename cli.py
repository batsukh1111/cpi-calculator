#!/usr/bin/env python3
"""
CPI тооцооллын командын мөр.

Жишээ:
  python cli.py calculate -i "cpi calculation 2023=100.xlsx" -o output/
  python cli.py calculate -i file.xlsx --aimag 01 20 --json
  python cli.py validate -i file.xlsx
"""

from __future__ import annotations

import argparse
import sys
import time
from pathlib import Path

# allow running without install
sys.path.insert(0, str(Path(__file__).resolve().parent))

from cpi.loader import load_workbook_data
from cpi.engine import CPICalculator, inflation_yoy
from cpi.export import export_excel, export_json, export_overall_csv
from cpi.special_groups import compute_ub_and_national_special
from cpi.regions import compute_all_regions
from cpi.period import parse_period, latest_period, three_way_changes, period_index
from cpi.publication import publish_for_period, generate_publication
import json as _json


def _latest_month_index(result) -> int:
    last_i = len(result.months) - 1
    for i in range(len(result.months) - 1, -1, -1):
        vals = [result.regions[c].overall[i] for c in result.regions]
        if any(abs(v - 100.0) > 1e-6 for v in vals) or abs(
            result.national.overall[i] - 100.0
        ) > 1e-6:
            return i
    return last_i


def cmd_calculate(args: argparse.Namespace) -> int:
    excel = Path(args.input)
    if not excel.exists():
        print(f"Файл олдсонгүй: {excel}", file=sys.stderr)
        return 1

    out = Path(args.output)
    out.mkdir(parents=True, exist_ok=True)

    aimag_codes = args.aimag
    print(f"Уншиж байна: {excel}")
    t0 = time.perf_counter()
    data = load_workbook_data(excel, aimag_codes=aimag_codes)
    print(
        f"  Аймаг: {len(data.regions)}, сар: {len(data.months)}, "
        f"мөр: {len(data.structure)} ({time.perf_counter()-t0:.1f}s)"
    )

    print("Тооцоолж байна...")
    t1 = time.perf_counter()
    calc = CPICalculator(data)
    result = calc.calculate()
    print(f"  Дууслаа ({time.perf_counter()-t1:.1f}s)")

    last_i = _latest_month_index(result)
    period = result.months[last_i]["period"]
    overall = result.national.overall[last_i]
    yoy = inflation_yoy(result.national.overall)
    yoy_v = yoy[last_i]
    print()
    print(f"=== Улсын ерөнхий индекс ({period}) ===")
    print(f"  Индекс (2023=100): {overall:.2f}")
    if yoy_v is not None:
        print(f"  Жилийн инфляц:     {yoy_v:.2f}%")
    print()
    print("Аймаг / нийслэл (сүүлийн сар):")
    for code in sorted(result.regions.keys()):
        rr = result.regions[code]
        print(
            f"  {code} {rr.name:14s}  {rr.overall[last_i]:8.2f}  "
            f"жин={rr.weights.get(8,0):.3f}"
        )

    # Special product groups (UB + national) + geographic regions
    special = None
    regional = None
    if "20" in result.regions:
        special = compute_ub_and_national_special(result)
        print()
        print(f"=== Тусгай барааны бүлэг ({period}) ===")
        print(f"{'Бүлэг':32s} {'УБ индекс':>10s} {'УБ жил%':>8s} {'Улс индекс':>10s} {'Улс жил%':>8s}")
        for key, g_ub in special["ulaanbaatar"].items():
            g_nat = special["national"][key]
            y_ub = g_ub["yoy"][last_i]
            y_nat = g_nat["yoy"][last_i]
            print(
                f"{g_ub['label_mn'][:32]:32s} "
                f"{g_ub['indices'][last_i]:10.2f} "
                f"{(f'{y_ub:.2f}' if y_ub is not None else '—'):>8s} "
                f"{g_nat['indices'][last_i]:10.2f} "
                f"{(f'{y_nat:.2f}' if y_nat is not None else '—'):>8s}"
            )

    if len(result.regions) >= 2:
        regional = compute_all_regions(result)
        print()
        print(f"=== Бүс (уламжлалт) — ерөнхий индекс ({period}) ===")
        for name, agg in regional["traditional"].items():
            y = inflation_yoy(agg["overall"])[last_i]
            yoy_s = f"{y:.2f}" if y is not None else "—"
            print(f"  {name:18s}  {agg['overall'][last_i]:8.2f}  жил%={yoy_s}")
        print(f"=== Бүс (шинэ) — ерөнхий индекс ({period}) ===")
        for name, agg in regional["new"].items():
            y = inflation_yoy(agg["overall"])[last_i]
            yoy_s = f"{y:.2f}" if y is not None else "—"
            print(f"  {name:18s}  {agg['overall'][last_i]:8.2f}  жил%={yoy_s}")

    if args.excel or not args.json_only:
        xlsx_path = out / "cpi_result.xlsx"
        export_excel(result, xlsx_path)
        print(f"\nExcel: {xlsx_path}")

    if args.json or args.json_only:
        json_path = out / "cpi_result.json"
        export_json(result, json_path, compact=not args.full_json)
        print(f"JSON:  {json_path}")
        if special:
            sp_path = out / "cpi_special_groups.json"
            # serializable dump
            sp_path.write_text(
                _json.dumps(special, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            print(f"JSON:  {sp_path}")
        if regional:
            # simplify regional for json
            reg_json = {}
            for scheme, regions in regional.items():
                reg_json[scheme] = {
                    name: {
                        "codes": agg["codes"],
                        "weight_total": agg["weight_total"],
                        "overall": agg["overall"],
                        "major": {
                            str(rid): {
                                "name": info["name"],
                                "weight": info["weight"],
                                "indices": info["indices"],
                            }
                            for rid, info in agg["rows"].items()
                        },
                    }
                    for name, agg in regions.items()
                }
            rp = out / "cpi_regions.json"
            rp.write_text(
                _json.dumps(reg_json, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            print(f"JSON:  {rp}")

    csv_path = out / "cpi_overall.csv"
    export_overall_csv(result, csv_path)
    print(f"CSV:   {csv_path}")

    return 0


def cmd_publish(args: argparse.Namespace) -> int:
    """Хугацаа оруулаад table 1–11 нийтлэлийн Excel гаргах."""
    excel = Path(args.input)
    if not excel.exists():
        print(f"Файл олдсонгүй: {excel}", file=sys.stderr)
        return 1

    template = Path(args.template) if args.template else Path(
        r"C:\Users\batsukh\Desktop\National_202607_2023.xlsx"
    )
    if not template.exists():
        print(f"Загвар файл олдсонгүй: {template}", file=sys.stderr)
        print("  --template замаар National_*.xlsx өгнө үү.")
        return 1

    out_dir = Path(args.output)
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"Уншиж байна: {excel}")
    t0 = time.perf_counter()
    data = load_workbook_data(excel)
    result = CPICalculator(data).calculate()
    print(f"  Тооцоо дууслаа ({time.perf_counter()-t0:.1f}s)")

    if args.period:
        period = parse_period(args.period)
    else:
        period = latest_period(result.months, result.national.overall)
        print(f"  Хугацаа автомат: {period.key}")

    t = period_index(result.months, period)
    ch = three_way_changes(result.national.overall, t)
    print()
    print(f"=== Улс — {period.title_mn()} ({period.roman}) ===")
    print(f"  Индекс (2023=100): {ch['index']:.2f}" if ch["index"] is not None else "  Индекс: —")
    if ch["yoy"] is not None:
        print(f"  Өмнөх оны мөн үе:     {ch['yoy']:+.1f}%")
    if ch["ytd"] is not None:
        print(f"  Өмнөх оны эцэс:       {ch['ytd']:+.1f}%")
    if ch["mom"] is not None:
        print(f"  Өмнөх сар:            {ch['mom']:+.1f}%")

    out_path = out_dir / f"National_{period.yyyymm}_2023.xlsx"
    print()
    print(f"Нийтлэлийн хүснэгт (table 1–11) үүсгэж байна...")
    print(f"  Загвар: {template}")
    generate_publication(
        result,
        data,
        period,
        template,
        out_path,
        tables_only=not args.keep_all_sheets,
    )
    print(f"  Хадгаллаа: {out_path}")

    # also write comparison summary csv
    import csv

    csv_path = out_dir / f"comparison_{period.yyyymm}.csv"
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "code",
                "name",
                "index",
                "yoy_pct",
                "ytd_pct",
                "mom_pct",
            ]
        )
        chn = three_way_changes(result.national.overall, t)
        w.writerow(
            [
                "00",
                "Улс",
                chn["index"],
                chn["yoy"],
                chn["ytd"],
                chn["mom"],
            ]
        )
        for code in sorted(result.regions.keys()):
            rr = result.regions[code]
            c = three_way_changes(rr.overall, t)
            w.writerow([code, rr.name, c["index"], c["yoy"], c["ytd"], c["mom"]])
    print(f"  Харьцуулалт CSV: {csv_path}")
    return 0


def cmd_validate(args: argparse.Namespace) -> int:
    """Excel-ийн data_only утгатай харьцуулах."""
    import openpyxl

    excel = Path(args.input)
    print(f"Тооцоолж байна: {excel}")
    data = load_workbook_data(excel)
    result = CPICalculator(data).calculate()

    print("Excel-ийн хадгалагдсан утгуудыг уншиж байна (data_only)...")
    wb = openpyxl.load_workbook(excel, data_only=True, read_only=True)

    def col_for_month(mi: int) -> int:
        return 11 + mi  # K=11

    max_check_months = min(24, len(result.months))  # first 2 years
    check_rows = [8, 9, 10, 11, 12, 13, 18, 209, 316]

    total = 0
    mismatches = 0
    max_diff = 0.0

    for code in ["01", "20"]:
        if code not in wb.sheetnames or code not in result.regions:
            continue
        ws = wb[code]
        rr = result.regions[code]
        print(f"\n--- Sheet {code} ({rr.name}) ---")
        for row in check_rows:
            for mi in range(max_check_months):
                excel_val = ws.cell(row, col_for_month(mi)).value
                if excel_val is None:
                    continue
                try:
                    excel_f = float(excel_val)
                except (TypeError, ValueError):
                    continue
                our = rr.indices[row][mi]
                diff = abs(our - excel_f)
                total += 1
                max_diff = max(max_diff, diff)
                if diff > args.tol:
                    mismatches += 1
                    if mismatches <= 15:
                        print(
                            f"  DIFF {code} r{row} m{result.months[mi]['period']}: "
                            f"ours={our:.6f} excel={excel_f:.6f} Δ={diff:.6f}"
                        )

    # National
    if "base index national" in wb.sheetnames:
        ws = wb["base index national"]
        print("\n--- National ---")
        for row in check_rows:
            for mi in range(max_check_months):
                excel_val = ws.cell(row, col_for_month(mi)).value
                if excel_val is None:
                    continue
                try:
                    excel_f = float(excel_val)
                except (TypeError, ValueError):
                    continue
                our = result.national.indices[row][mi]
                diff = abs(our - excel_f)
                total += 1
                max_diff = max(max_diff, diff)
                if diff > args.tol:
                    mismatches += 1
                    if mismatches <= 25:
                        print(
                            f"  DIFF NAT r{row} m{result.months[mi]['period']}: "
                            f"ours={our:.6f} excel={excel_f:.6f} Δ={diff:.6f}"
                        )

    wb.close()
    print(f"\nШалгасан: {total}, зөрүү (>{args.tol}): {mismatches}, max|Δ|={max_diff:.8f}")
    if mismatches == 0:
        print("OK — Excel-тэй таарч байна.")
        return 0
    print("АНХААР: зөрүү олдлоо.")
    return 1


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        description="Хэрэглээний үнийн индекс (CPI) тооцоолуур — 2023=100"
    )
    sub = p.add_subparsers(dest="cmd", required=True)

    c = sub.add_parser("calculate", help="Индекс тооцоод файлд хадгалах")
    c.add_argument("-i", "--input", required=True, help="cpi calculation Excel файл")
    c.add_argument("-o", "--output", default="output", help="гаралтын хавтас")
    c.add_argument(
        "--aimag",
        nargs="+",
        default=None,
        help="Зөвхөн эдгээр аймгууд (жишээ: 01 20). Хоосон=бүгд",
    )
    c.add_argument("--json", action="store_true", help="JSON экспорт нэмэх")
    c.add_argument("--json-only", action="store_true", help="Зөвхөн JSON")
    c.add_argument("--full-json", action="store_true", help="Бүх мөрийн индекс JSON-д")
    c.add_argument("--excel", action="store_true", default=True)
    c.set_defaults(func=cmd_calculate)

    v = sub.add_parser("validate", help="Excel-ийн утгатай харьцуулах")
    v.add_argument("-i", "--input", required=True)
    v.add_argument("--tol", type=float, default=0.01, help="зөвшөөрөгдөх зөрүү")
    v.set_defaults(func=cmd_validate)

    pub = sub.add_parser(
        "publish",
        help="Хугацаагаар YoY/YTD/MoM тооцоод table 1–11 гаргах",
    )
    pub.add_argument(
        "-i",
        "--input",
        required=True,
        help="cpi calculation 2023=100.xlsx",
    )
    pub.add_argument(
        "-p",
        "--period",
        default=None,
        help="Хугацаа: 2026-06 эсвэл 2026.06 (хоосон=сүүлийн сар)",
    )
    pub.add_argument(
        "-t",
        "--template",
        default=None,
        help="National_*.xlsx загвар (table 1–11)",
    )
    pub.add_argument("-o", "--output", default="output", help="гаралтын хавтас")
    pub.add_argument(
        "--keep-all-sheets",
        action="store_true",
        help="Загварын бүх sheet үлдээх (default: зөвхөн table 1–11)",
    )
    pub.set_defaults(func=cmd_publish)

    args = p.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
