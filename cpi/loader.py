"""
Excel файлыг уншиж жин, үнэ, бүтцийг гаргана.

Файлын бүтэц (cpi calculation 2023=100.xlsx):
  - Sheet 01–22: аймаг / нийслэл
  - Sheet «base index national»: улсын жигнэсэн индекс
  - Мөр 8–723: индексийн шатлал (COICOP)
  - Мөр 728–1443: сарын дундаж үнэ (index row + 720 = price row)
  - Багана H: жин, I: харьцангуй жин, J: суурь=100 / суурь үнэ
  - Багана K+: сар бүрийн индекс / үнэ
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

# Багцын data/ хавтас
_DATA_DIR = Path(__file__).resolve().parent.parent / "data"
PRICE_OFFSET = 720
INDEX_START = 8
INDEX_END = 723
PRICE_START = INDEX_START + PRICE_OFFSET  # 728 labels, items from 733
MONTH_COL0 = 10  # column K (0-based index 10)


@dataclass
class Node:
    row: int
    name: str | None
    level: int | None
    item_no: int | None
    kind: str  # elementary | sumproduct | weighted_children
    children: list[int] = field(default_factory=list)
    price_row: int | None = None


@dataclass
class RegionData:
    code: str
    name: str
    weights: dict[int, float]  # row -> absolute weight H
    prices: dict[int, list[float | None]]  # price_row -> [month prices]


@dataclass
class WorkbookData:
    structure: list[Node]
    nodes_by_row: dict[int, Node]
    months: list[dict[str, Any]]
    aimags: dict[str, str]
    regions: dict[str, RegionData]
    base_year: int = 2023
    base_month_count: int = 12  # 2023.01–2023.12 for base average


def _load_json(name: str) -> Any:
    path = _DATA_DIR / name
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def load_structure() -> tuple[list[Node], dict[int, Node]]:
    raw = _load_json("structure.json")
    nodes = [
        Node(
            row=s["row"],
            name=s.get("name"),
            level=s.get("level"),
            item_no=s.get("item_no"),
            kind=s["kind"],
            children=list(s.get("children") or []),
            price_row=s.get("price_row"),
        )
        for s in raw
    ]
    by_row = {n.row: n for n in nodes}
    return nodes, by_row


def load_aimags() -> dict[str, str]:
    return _load_json("aimags.json")


def load_months() -> list[dict[str, Any]]:
    return _load_json("months.json")


def detect_months_from_sheet(ws, header_row: int = 7) -> list[dict[str, Any]]:
    """
    Sheet-ийн header мөрөөс сарын багануудыг автоматаар уншина.
    «2023.01/ 2023», «2026.07/ 2023» гэх мэт.
    Шинэ сар (ж.нь 7-р сар) нэмэгдэхэд months.json шинэчлэх шаардлагагүй.
    """
    months: list[dict[str, Any]] = []
    # Read a wide row (up to col 120)
    for row in ws.iter_rows(
        min_row=header_row, max_row=header_row, min_col=1, max_col=120, values_only=False
    ):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            s = str(v).strip()
            m = re.match(r"(\d{4})\.(\d{2})", s)
            if not m:
                continue
            # 0-based column index for openpyxl values_only rows
            col0 = cell.column - 1  # openpyxl column is 1-based
            months.append(
                {
                    "col": col0,
                    "label": s,
                    "period": f"{m.group(1)}-{m.group(2)}",
                }
            )
    # Keep only contiguous price/index months starting near K (col 10+)
    # Prefer labels that look like "YYYY.MM/ YYYY"
    months = [x for x in months if x["col"] >= MONTH_COL0]
    months.sort(key=lambda x: x["col"])
    return months


def _to_float(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (v != v):  # NaN
            return None
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if not s or s.startswith("#"):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _read_region_sheet(
    ws,
    code: str,
    name: str,
    months: list[dict[str, Any]],
    index_rows: list[int],
    price_rows: list[int],
) -> RegionData:
    """Унших: H багана (жин) index мөрүүдээс, үнэ price мөрүүдээс."""
    weights: dict[int, float] = {}
    prices: dict[int, list[float | None]] = {}
    n_months = len(months)
    # absolute max col needed (1-based openpyxl)
    max_col0 = max((m["col"] for m in months), default=MONTH_COL0) 
    max_col = max_col0 + 1  # openpyxl max_col is 1-based inclusive when used carefully
    # values_only rows are 0-indexed by position from min_col=1 → index = col-1

    # Index section: weights in col H (index 7)
    for i, row in enumerate(
        ws.iter_rows(
            min_row=INDEX_START,
            max_row=INDEX_END,
            min_col=1,
            max_col=8,
            values_only=True,
        ),
        INDEX_START,
    ):
        h = _to_float(row[7] if len(row) > 7 else None)
        weights[i] = h if h is not None else 0.0

    # Price section: use actual column indices from months detection
    read_max = max_col0 + 1
    for i, row in enumerate(
        ws.iter_rows(
            min_row=PRICE_START,
            max_row=PRICE_START + (INDEX_END - INDEX_START),
            min_col=1,
            max_col=read_max,
            values_only=True,
        ),
        PRICE_START,
    ):
        month_vals = []
        for m in months:
            col_i = m["col"]  # 0-based in values_only row starting min_col=1
            val = row[col_i] if col_i < len(row) else None
            month_vals.append(_to_float(val))
        prices[i] = month_vals

    return RegionData(code=code, name=name, weights=weights, prices=prices)


def load_workbook_data(
    excel_path: str | Path,
    aimag_codes: list[str] | None = None,
) -> WorkbookData:
    """
    Бүрэн Excel файлыг уншина.

    Parameters
    ----------
    excel_path : path to cpi calculation 2023=100.xlsx
    aimag_codes : None = бүх 01–22
    """
    excel_path = Path(excel_path)
    structure, by_row = load_structure()
    aimags = load_aimags()

    if aimag_codes is None:
        aimag_codes = sorted(aimags.keys())

    index_rows = [n.row for n in structure]
    price_rows = [
        n.price_row for n in structure if n.kind == "elementary" and n.price_row
    ]

    wb = openpyxl.load_workbook(excel_path, data_only=False, read_only=True)
    regions: dict[str, RegionData] = {}

    try:
        # Саруудыг Excel header-ээс автоматаар (шинэ сар нэмэгдвэл аяндаа)
        probe = None
        for code in ("20", "01") + tuple(aimag_codes):
            if code in wb.sheetnames:
                probe = code
                break
        if probe is None:
            raise KeyError(f"Аймгийн sheet олдсонгүй: {excel_path}")
        months = detect_months_from_sheet(wb[probe])
        if not months:
            months = load_months()  # fallback
        # Persist for reference (optional)
        try:
            with open(_DATA_DIR / "months.json", "w", encoding="utf-8") as f:
                json.dump(months, f, ensure_ascii=False, indent=2)
        except OSError:
            pass

        for code in aimag_codes:
            if code not in wb.sheetnames:
                raise KeyError(f"Sheet '{code}' олдсонгүй: {excel_path}")
            name = aimags.get(code, code)
            # 20-ийн нэр файлд «Нийслэл»
            if code == "20":
                name = "Улаанбаатар"
            ws = wb[code]
            regions[code] = _read_region_sheet(
                ws, code, name, months, index_rows, price_rows
            )
    finally:
        wb.close()

    return WorkbookData(
        structure=structure,
        nodes_by_row=by_row,
        months=months,
        aimags=aimags,
        regions=regions,
        base_year=2023,
        base_month_count=12,
    )


def parse_structure_from_sheet(ws) -> list[dict]:
    """Шаардлагатай бол Excel-ээс бүтцийг дахин гаргана (maintenance)."""
    structure = []
    for i, row in enumerate(
        ws.iter_rows(min_row=INDEX_START, max_row=INDEX_END, max_col=11, values_only=True),
        INDEX_START,
    ):
        cells = list(row) + [None] * 11
        A, B, C, D, E, F, G, H, I, J, K = cells[:11]
        name = None
        level = None
        for val, lev in [(A, 0), (B, 1), (C, 2), (D, 3)]:
            if isinstance(val, str) and val.strip():
                name = val.strip()
                level = lev
                break
        if name is None and G:
            name = str(G).strip()
            level = 4
        k = K if isinstance(K, str) else None
        children: list[int] = []
        kind = "empty"
        price_row = None
        if k:
            if "SUMPRODUCT" in k:
                kind = "sumproduct"
                m = re.findall(r"\$?I\$?(\d+):\$?I\$?(\d+)", k)
                if m:
                    a, b = int(m[0][0]), int(m[0][1])
                    children = list(range(a, b + 1))
                else:
                    m2 = re.findall(r"SUMPRODUCT\(\$?I\$?(\d+)", k)
                    if m2:
                        children = [int(m2[0])]
            elif re.search(r"[A-Z](7\d{2}|8\d{2}|9\d{2}|1[0-4]\d{2})", k) and "IF" in k:
                kind = "elementary"
                refs = [int(x) for x in re.findall(r"[A-Z](\d{3,4})", k) if int(x) >= 720]
                price_row = refs[0] if refs else i + PRICE_OFFSET
            else:
                refs = re.findall(r"\$?I\$?(\d+)\s*\*\s*[A-Z](\d+)", k)
                if refs:
                    kind = "weighted_children"
                    children = [int(a) for a, _ in refs]
                else:
                    kind = "other"
        structure.append(
            {
                "row": i,
                "name": name,
                "level": level,
                "item_no": int(F) if isinstance(F, (int, float)) else None,
                "weight_template": float(H) if isinstance(H, (int, float)) else 0.0,
                "kind": kind,
                "children": children,
                "price_row": price_row,
            }
        )
    return structure
