"""Үр дүнг Excel / CSV / JSON-д экспортлох."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .engine import CPIResult, inflation_mom, inflation_yoy


MAJOR_ROWS = [8, 9, 209, 229, 316, 377, 461, 505, 554, 581, 628, 647, 663, 684]


def _period_labels(result: CPIResult) -> list[str]:
    return [m["period"] for m in result.months]


def summary_table(
    result: CPIResult,
    scope: str = "national",
    code: str | None = None,
) -> list[dict[str, Any]]:
    """Ерөнхий + 13 бүлгийн индекс хүснэгт."""
    periods = _period_labels(result)
    if scope == "national":
        indices = result.national.indices
        weights = result.national.weights
        name = "Улс"
    else:
        rr = result.regions[code or "20"]
        indices = rr.indices
        weights = rr.weights
        name = rr.name

    by_row = {n.row: n for n in result.structure}
    rows = []
    for r in MAJOR_ROWS:
        node = by_row.get(r)
        series = indices.get(r, [100.0] * len(periods))
        yoy = inflation_yoy(series)
        mom = inflation_mom(series)
        rows.append(
            {
                "scope": name,
                "row": r,
                "name": node.name if node else str(r),
                "weight": weights.get(r, 0.0),
                "indices": {periods[i]: series[i] for i in range(len(periods))},
                "yoy": {periods[i]: yoy[i] for i in range(len(periods))},
                "mom": {periods[i]: mom[i] for i in range(len(periods))},
            }
        )
    return rows


def export_json(result: CPIResult, path: str | Path, compact: bool = True) -> None:
    path = Path(path)
    periods = _period_labels(result)
    payload: dict[str, Any] = {
        "base_year": result.base_year,
        "months": periods,
        "national": {
            "overall": result.national.overall,
            "major_groups": {},
        },
        "regions": {},
    }
    by_row = {n.row: n for n in result.structure}
    for r in MAJOR_ROWS:
        node = by_row.get(r)
        label = node.name if node else str(r)
        payload["national"]["major_groups"][label] = {
            "weight": result.national.weights.get(r, 0.0),
            "indices": result.national.indices.get(r, []),
        }

    for code, rr in result.regions.items():
        reg: dict[str, Any] = {
            "name": rr.name,
            "overall": rr.overall,
            "weight_total": rr.weights.get(8, 0.0),
            "major_groups": {},
        }
        if not compact:
            reg["all_indices"] = {
                str(row): vals for row, vals in rr.indices.items()
            }
        for r in MAJOR_ROWS:
            node = by_row.get(r)
            label = node.name if node else str(r)
            reg["major_groups"][label] = {
                "weight": rr.weights.get(r, 0.0),
                "indices": rr.indices.get(r, []),
            }
        payload["regions"][code] = reg

    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def _styles():
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    return header_font, header_fill, thin


def export_special_and_regions(
    result: CPIResult,
    path: str | Path,
    special: dict | None = None,
    regional: dict | None = None,
) -> None:
    """
    Тусгай барааны бүлэг (УБ + улс) болон бүсийн индексийг Excel-д бичнэ.
    path — шинэ файл эсвэл одоо байгаа workbook-д sheet нэмэхэд ашиглана.
    """
    from .special_groups import compute_ub_and_national_special
    from .regions import compute_all_regions

    path = Path(path)
    periods = _period_labels(result)
    header_font, header_fill, thin = _styles()

    if special is None:
        try:
            special = compute_ub_and_national_special(result)
        except KeyError:
            special = None
    if regional is None:
        regional = compute_all_regions(result)

    if path.exists():
        from openpyxl import load_workbook

        wb = load_workbook(path)
    else:
        wb = Workbook()
        # drop default if empty new
        if wb.active.title == "Sheet" and wb.active.max_row == 1:
            pass

    def write_special_sheet(ws, title: str, groups: dict):
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = "Эх сурвалж: Бүлэг.xlsx → барааны бүлэг | суурь 2023=100"
        headers = ["Түлхүүр", "Бүлэг", "Жин", "Бараа тоо"] + periods
        for c, h in enumerate(headers, 1):
            cell = ws.cell(3, c, h)
            cell.font = header_font
            cell.fill = header_fill
        # stable order matching file columns
        order = [
            "domestic",
            "domestic_ex_meat",
            "import",
            "meat",
            "fuel",
            "import_ex_fuel",
            "goods",
            "services",
            "core",
            "food",
            "non_food",
            "electricity_fuel",
        ]
        r = 4
        for key in order:
            if key not in groups:
                continue
            g = groups[key]
            ws.cell(r, 1, key)
            ws.cell(r, 2, g["label_mn"])
            ws.cell(r, 3, round(g["weight"], 6))
            ws.cell(r, 4, g["n_items"])
            for m_i, val in enumerate(g["indices"]):
                cell = ws.cell(r, 5 + m_i, round(val, 4))
                cell.number_format = "0.00"
                cell.border = thin
            r += 1
        # YoY block
        r += 1
        ws.cell(r, 1, "Жилийн өөрчлөлт (%)").font = Font(bold=True)
        r += 1
        for c, h in enumerate(["Түлхүүр", "Бүлэг"] + periods, 1):
            cell = ws.cell(r, c, h)
            cell.font = header_font
            cell.fill = header_fill
        r += 1
        start_yoy = r
        for key in order:
            if key not in groups:
                continue
            g = groups[key]
            ws.cell(r, 1, key)
            ws.cell(r, 2, g["label_mn"])
            for m_i, val in enumerate(g["yoy"]):
                cell = ws.cell(
                    r, 3 + m_i, round(val, 2) if val is not None else None
                )
                cell.number_format = "0.00"
            r += 1
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 10

    if special:
        for sheet_name, scope_key, title in [
            ("Тусгай_УБ", "ulaanbaatar", "Улаанбаатар — тусгай барааны бүлгийн индекс"),
            ("Тусгай_Улс", "national", "Улс — тусгай барааны бүлгийн индекс"),
        ]:
            if scope_key not in special:
                continue
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(sheet_name)
            write_special_sheet(ws, title, special[scope_key])

    def write_region_scheme(ws, title: str, scheme: dict):
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = "Аймгуудын жингээр жигнэсэн | суурь 2023=100"
        headers = ["Бүс", "Аймгууд", "Жин"] + periods
        for c, h in enumerate(headers, 1):
            cell = ws.cell(3, c, h)
            cell.font = header_font
            cell.fill = header_fill
        r = 4
        for name, agg in scheme.items():
            codes = ",".join(agg["codes"])
            ws.cell(r, 1, name)
            ws.cell(r, 2, codes)
            ws.cell(r, 3, round(agg["weight_total"], 6))
            for m_i, val in enumerate(agg["overall"]):
                cell = ws.cell(r, 4 + m_i, round(val, 4))
                cell.number_format = "0.00"
            r += 1
        # major groups per region (overall-style table for food etc.)
        r += 2
        ws.cell(r, 1, "COICOP бүлгүүд (сүүлийн сарын индекс)").font = Font(bold=True)
        r += 1
        last_i = len(periods) - 1
        for i in range(len(periods) - 1, -1, -1):
            # use last period index position
            last_i = i
            break
        headers2 = ["Бүс", "Бүлэг", "Жин", f"Индекс {periods[last_i]}", "Жилийн %"]
        for c, h in enumerate(headers2, 1):
            cell = ws.cell(r, c, h)
            cell.font = header_font
            cell.fill = header_fill
        r += 1
        for name, agg in scheme.items():
            for row_id, info in agg["rows"].items():
                yoy = info["yoy"][last_i]
                ws.cell(r, 1, name)
                ws.cell(r, 2, info["name"])
                ws.cell(r, 3, round(info["weight"], 6))
                ws.cell(r, 4, round(info["indices"][last_i], 4)).number_format = "0.00"
                ws.cell(
                    r, 5, round(yoy, 2) if yoy is not None else None
                ).number_format = "0.00"
                r += 1
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 12

    if regional:
        for sheet_name, scheme_key, title in [
            ("Бүс_уламжлалт", "traditional", "Уламжлалт бүс — ерөнхий индекс"),
            ("Бүс_шинэ", "new", "Шинэ бүс — ерөнхий индекс"),
        ]:
            if scheme_key not in regional:
                continue
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(sheet_name)
            write_region_scheme(ws, title, regional[scheme_key])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def export_excel(result: CPIResult, path: str | Path) -> None:
    """Хураангуй Excel: Улс + аймаг бүрийн ерөнхий индекс, бүлгүүд + тусгай/бүс."""
    path = Path(path)
    wb = Workbook()
    periods = _period_labels(result)
    header_font, header_fill, thin = _styles()

    def write_sheet(ws, title: str, weights, indices, by_row):
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Суурь жил: {result.base_year}=100"
        headers = ["Мөр", "Нэр", "Жин"] + periods
        for c, h in enumerate(headers, 1):
            cell = ws.cell(3, c, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, horizontal="center")
        for r_i, row in enumerate(MAJOR_ROWS, 4):
            node = by_row.get(row)
            ws.cell(r_i, 1, row)
            ws.cell(r_i, 2, node.name if node else "")
            ws.cell(r_i, 3, round(weights.get(row, 0.0), 6))
            series = indices.get(row, [])
            for m_i, val in enumerate(series):
                cell = ws.cell(r_i, 4 + m_i, round(val, 4) if val is not None else None)
                cell.number_format = "0.00"
                cell.border = thin
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 12
        for m_i in range(len(periods)):
            ws.column_dimensions[get_column_letter(4 + m_i)].width = 11

    by_row = {n.row: n for n in result.structure}

    # National
    ws = wb.active
    ws.title = "Улс"
    write_sheet(
        ws,
        "Улсын хэрэглээний үнийн индекс (2023=100)",
        result.national.weights,
        result.national.indices,
        by_row,
    )

    # Overall comparison sheet
    ws_all = wb.create_sheet("Аймаг_ерөнхий")
    ws_all["A1"] = "Аймаг / бүс бүрийн ЕРӨНХИЙ ИНДЕКС"
    ws_all["A1"].font = Font(bold=True, size=14)
    headers = ["Код", "Нэр", "Жин"] + periods
    for c, h in enumerate(headers, 1):
        cell = ws_all.cell(3, c, h)
        cell.font = header_font
        cell.fill = header_fill
    # national first
    ws_all.cell(4, 1, "00")
    ws_all.cell(4, 2, "Улс")
    ws_all.cell(4, 3, round(result.national.weights.get(8, 0.0), 6))
    for m_i, val in enumerate(result.national.overall):
        ws_all.cell(4, 4 + m_i, round(val, 4)).number_format = "0.00"
    r = 5
    for code in sorted(result.regions.keys()):
        rr = result.regions[code]
        ws_all.cell(r, 1, code)
        ws_all.cell(r, 2, rr.name)
        ws_all.cell(r, 3, round(rr.weights.get(8, 0.0), 6))
        for m_i, val in enumerate(rr.overall):
            ws_all.cell(r, 4 + m_i, round(val, 4)).number_format = "0.00"
        r += 1
    ws_all.column_dimensions["A"].width = 8
    ws_all.column_dimensions["B"].width = 18
    ws_all.column_dimensions["C"].width = 12

    # YoY inflation national
    ws_inf = wb.create_sheet("Инфляц_Улс")
    ws_inf["A1"] = "Улсын инфляц (%), 12 сарын өөрчлөлт"
    ws_inf["A1"].font = Font(bold=True, size=14)
    headers = ["Бүлэг"] + periods
    for c, h in enumerate(headers, 1):
        cell = ws_inf.cell(3, c, h)
        cell.font = header_font
        cell.fill = header_fill
    for r_i, row in enumerate(MAJOR_ROWS, 4):
        node = by_row.get(row)
        series = result.national.indices.get(row, [])
        yoy = inflation_yoy(series)
        ws_inf.cell(r_i, 1, node.name if node else str(row))
        for m_i, val in enumerate(yoy):
            cell = ws_inf.cell(r_i, 2 + m_i, round(val, 2) if val is not None else None)
            cell.number_format = "0.00"
    ws_inf.column_dimensions["A"].width = 55

    # Per-region sheets (major groups only)
    for code in sorted(result.regions.keys()):
        rr = result.regions[code]
        title = f"{code}_{rr.name}"[:31]
        ws = wb.create_sheet(title)
        write_sheet(
            ws,
            f"{rr.name} — хэрэглээний үнийн индекс (2023=100)",
            rr.weights,
            rr.indices,
            by_row,
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

    # Append special product groups (UB + national) and geographic regions
    try:
        export_special_and_regions(result, path)
    except Exception as e:
        # Core export succeeded; special sheets optional
        print(f"Анхааруулга: тусгай/бүсийн sheet бичихэд алдаа: {e}")


def export_overall_csv(result: CPIResult, path: str | Path) -> None:
    import csv

    path = Path(path)
    periods = _period_labels(result)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["code", "name", "weight"] + periods)
        w.writerow(
            ["00", "Улс", result.national.weights.get(8, 0.0)]
            + [round(x, 6) for x in result.national.overall]
        )
        for code in sorted(result.regions.keys()):
            rr = result.regions[code]
            w.writerow(
                [code, rr.name, rr.weights.get(8, 0.0)]
                + [round(x, 6) for x in rr.overall]
            )
