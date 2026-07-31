"""
Олон нийтэд тараах хүснэгт (table 1–11) — National_YYYYMM_2023 формат.

Хэрэглэгч хугацаа (ж.нь 2026-06) оруулахад:
  - өмнөх оны мөн үе (YoY)
  - өмнөх оны эцэс (YTD / XII)
  - өмнөх сар (MoM)
харьцуулалтыг тооцоод table 1–11-ийг автоматаар гаргана.
"""

from __future__ import annotations

import re
import shutil
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .engine import CPIResult
from .period import Period, period_index, three_way_changes, parse_period, latest_period
from .regions import compute_all_regions, aggregate_region_codes
from .special_groups import load_special_groups, special_group_index
from .export import MAJOR_ROWS

# Traditional region codes (matches National file west/khangai/center/east)
REGION_SHEETS = {
    "west": ["02", "05", "09", "15", "16"],
    "khangai": ["01", "03", "04", "21", "10", "17"],
    "center": ["22", "19", "06", "08", "11", "13", "14"],
    "east": ["07", "12", "18"],
}

# Special group overall series key -> sheet name in National file
SPECIAL_SHEET_MAP = {
    # national
    ("national", "goods"): "baraa_nat",
    ("national", "services"): "uilchilgee_nat",
    ("national", "food"): "nat_huns",
    ("national", "non_food"): "nat_huns bus",
    # UB
    ("20", "goods"): "ubbaraa",
    ("20", "services"): "ubuilchilgee",
    ("20", "food"): "ubhuns",
    ("20", "non_food"): "ubhunsbus",
    # region specials
    ("west", "goods"): "west_baraa",
    ("west", "services"): "west_uil",
    ("west", "food"): "west_huns",
    ("west", "non_food"): "west_huns bus",
    ("khangai", "goods"): "khangai_baraa",
    ("khangai", "services"): "khangai_uil",
    ("khangai", "food"): "khangai_huns",
    ("khangai", "non_food"): "khangai_huns bus",
    ("center", "goods"): "center_baraa",
    ("center", "services"): "center_uil",
    ("center", "food"): "center_huns",
    ("center", "non_food"): "center_huns bus",
    ("east", "goods"): "east_baraa",
    ("east", "services"): "east_uil",
    ("east", "food"): "east_huns",
    ("east", "non_food"): "east_huns bus",
}

COMP_COL = {"BD": "yoy", "BE": "ytd", "BF": "mom"}


def _all_structure_rows(result: CPIResult) -> list[int]:
    return [n.row for n in result.structure]


def _build_region_full(result: CPIResult) -> dict[str, dict[int, list[float]]]:
    """Бүс бүрт бүх COICOP мөрийн индекс."""
    out = {}
    rows = _all_structure_rows(result)
    for name, codes in REGION_SHEETS.items():
        agg = aggregate_region_codes(result, codes, rows=rows)
        out[name] = {rid: info["indices"] for rid, info in agg["rows"].items()}
    return out


def _special_series_for_codes(
    result: CPIResult,
    codes: list[str],
    member_rows: list[int],
) -> list[float]:
    """Олон аймгийн special group: Σ H_a,i * I_a,i / Σ H_a,i."""
    n = len(result.months)
    total_w = 0.0
    # accumulate weight and weighted index sum per month
    w_by_item: dict[int, float] = {}
    for code in codes:
        if code not in result.regions:
            continue
        rr = result.regions[code]
        for r in member_rows:
            w = rr.weights.get(r, 0.0) or 0.0
            if w == 0:
                continue
            w_by_item[r] = w_by_item.get(r, 0.0) + w

    total_w = sum(w_by_item.values())
    if total_w == 0:
        return [100.0] * n

    out = []
    for t in range(n):
        s = 0.0
        for r, w in w_by_item.items():
            # national-style: weight is sum of aimag weights; index is weight-avg of aimag indices
            num = 0.0
            den = 0.0
            for code in codes:
                if code not in result.regions:
                    continue
                rr = result.regions[code]
                wa = rr.weights.get(r, 0.0) or 0.0
                if wa == 0 or r not in rr.indices:
                    continue
                num += wa * rr.indices[r][t]
                den += wa
            idx = (num / den) if den else 100.0
            s += w * idx
        out.append(s / total_w)
    return out


def build_lookup(
    result: CPIResult,
    period: Period,
) -> dict[str, dict[int, dict[str, float | None]]]:
    """
    sheet_name -> structure_row -> {yoy, ytd, mom, index}

    Special group sheets only use row 8 (overall of that group).
    """
    t = period_index(result.months, period)
    lookup: dict[str, dict[int, dict[str, float | None]]] = {}

    def put_series(sheet: str, row: int, series: list[float]):
        lookup.setdefault(sheet, {})[row] = three_way_changes(series, t)

    # National
    for row, series in result.national.indices.items():
        put_series("base index national", row, series)

    # Aimags
    for code, rr in result.regions.items():
        for row, series in rr.indices.items():
            put_series(code, row, series)

    # Geographic regions (full tree)
    reg_idx = _build_region_full(result)
    for sheet, by_row in reg_idx.items():
        for row, series in by_row.items():
            put_series(sheet, row, series)

    # Special groups: national, UB, regions
    groups_cfg = load_special_groups()
    groups = groups_cfg["groups"]
    keys_needed = ["goods", "services", "food", "non_food"]

    for gkey in keys_needed:
        members = groups.get(gkey, [])
        # national
        series, _ = special_group_index(
            members, result.national.weights, result.national.indices, len(result.months)
        )
        sheet = SPECIAL_SHEET_MAP.get(("national", gkey))
        if sheet:
            put_series(sheet, 8, series)

        # UB
        if "20" in result.regions:
            ub = result.regions["20"]
            series, _ = special_group_index(
                members, ub.weights, ub.indices, len(result.months)
            )
            sheet = SPECIAL_SHEET_MAP.get(("20", gkey))
            if sheet:
                put_series(sheet, 8, series)

        # regions
        for rname, codes in REGION_SHEETS.items():
            series = _special_series_for_codes(result, codes, members)
            sheet = SPECIAL_SHEET_MAP.get((rname, gkey))
            if sheet:
                put_series(sheet, 8, series)

    return lookup


def _resolve_cell(
    sheet: str,
    col: str,
    row: int,
    lookup: dict[str, dict[int, dict[str, float | None]]],
    price_lookup: dict[str, dict[int, dict[str, float | None]]] | None = None,
) -> float | None:
    """BD/BE/BF → yoy/ytd/mom; monthly price cols → price_lookup."""
    if col in COMP_COL:
        mode = COMP_COL[col]
        return lookup.get(sheet, {}).get(row, {}).get(mode)

    # Price columns (table 11): AD, AP, BB etc. — handled via price_lookup by period
    if price_lookup and sheet in price_lookup and row in price_lookup[sheet]:
        return price_lookup[sheet][row].get(col)

    return None


def _build_price_lookup(
    result: CPIResult,
    period: Period,
    years_back: int = 2,
) -> dict[str, dict[int, dict[str, float | None]]]:
    """
    Table 11: staple food prices for June of current and previous years.
    Keys by aimag code and price_row; subkey is synthetic 'Y0','Y1','Y2' for years.
    Actually we map Excel col of the template dynamically when filling.
    Here store: code -> price_row -> {period_key: price}
    """
    # result.regions[code]. prices are in RegionData — need from original
    # We only have indices in CPIResult. Prices must be passed separately.
    return {}


def build_price_lookup_from_data(
    data,
    period: Period,
) -> dict[tuple[str, int], dict[str, float | None]]:
    """
    (aimag_code, price_row) -> { '2024-06': price, '2025-06': price, '2026-06': price }
    for same calendar month across years.
    """
    out: dict[tuple[str, int], dict[str, float | None]] = {}
    months = data.months
    # find all June (or period.month) indices
    targets = []
    for y in range(period.year - 2, period.year + 1):
        pkey = f"{y:04d}-{period.month:02d}"
        for i, m in enumerate(months):
            if m["period"] == pkey:
                targets.append((pkey, i))
                break

    for code, region in data.regions.items():
        for prow, prices in region.prices.items():
            d = {}
            for pkey, mi in targets:
                if mi < len(prices):
                    d[pkey] = prices[mi]
                else:
                    d[pkey] = None
            out[(code, prow)] = d
    return out


def _update_period_text(value: Any, period: Period) -> Any:
    """Гарчгийн 2026 оны 6 / June 2026 / 2026 VI гэх мэтийг солих."""
    if not isinstance(value, str):
        return value
    s = value
    # Mongolian full
    s = re.sub(
        r"\d{4}\s*оны\s*\d{1,2}\s*дугаар\s*сард",
        period.title_mn(),
        s,
    )
    # English in Month Year
    s = re.sub(
        r"in\s+(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}",
        period.title_en(),
        s,
        flags=re.I,
    )
    return s


def _header_period_labels(period: Period) -> dict[str, str]:
    yoy_p = period.same_month_prev_year()
    ytd_p = period.end_prev_year()
    mom_p = period.prev_month()
    return {
        "cur": period.roman,  # 2026 VI
        "yoy": yoy_p.roman,  # 2025 VI
        "ytd": ytd_p.roman,  # 2025 XII
        "mom": mom_p.roman,  # 2026 V
    }


def generate_publication(
    result: CPIResult,
    data,
    period: Period | str,
    template_path: str | Path,
    output_path: str | Path,
    tables_only: bool = True,
) -> Path:
    """
    National template-ээс table 1–11-ийг хуулж, сонгосон хугацаагаар дүүргэнэ.
    """
    if isinstance(period, str):
        period = parse_period(period)

    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    t = period_index(result.months, period)
    lookup = build_lookup(result, period)
    prices = build_price_lookup_from_data(data, period)
    labels = _header_period_labels(period)

    # Copy template
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)

    table_names = [f"table {i}" for i in range(1, 12)]

    # Map Excel monthly col letter (in National file) -> period key for table 11
    # Col M=2023-01 ... BB=2026-06. We'll resolve by matching formula refs.

    for sn in table_names:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]

        # Update title / header text and period column headers
        for row in ws.iter_rows(min_row=1, max_row=8, max_col=min(ws.max_column or 20, 40)):
            for cell in row:
                if isinstance(cell.value, str) and not cell.value.startswith("="):
                    new = _update_period_text(cell.value, period)
                    # Replace roman period triplets in headers
                    if re.search(r"\d{4}\s+[IVX]+", str(cell.value)):
                        # sequential replacements carefully
                        pass
                    cell.value = new

        # Specific header cells for comparison tables (table 1, 6, 7)
        # Row 6-7 often have 2026 VI / 2025 VI / 2025 XII / 2026 V
        for r in range(1, 9):
            for c in range(1, min((ws.max_column or 15) + 1, 20)):
                cell = ws.cell(r, c)
                v = cell.value
                if not isinstance(v, str) or v.startswith("="):
                    continue
                # exact roman year-month headers
                if re.fullmatch(r"\d{4}\s+[IVX]+", v.strip()):
                    # determine role by column position in tables 1,6,7
                    pass

        # Force comparison headers on known layout
        if sn in ("table 1", "table 6", "table 7"):
            # table 1: E6=cur, F6=cur, G6=cur; E7=yoy, F7=ytd, G7=mom
            # table 6/7 similar but start at E3/E4
            header_map = {
                "table 1": [(6, 5, "cur"), (6, 6, "cur"), (6, 7, "cur"),
                            (7, 5, "yoy"), (7, 6, "ytd"), (7, 7, "mom")],
                "table 6": [(3, 5, "cur"), (3, 6, "cur"), (3, 7, "cur"),
                            (4, 5, "yoy"), (4, 6, "ytd"), (4, 7, "mom")],
                "table 7": [(3, 3, "cur"), (3, 4, "cur"), (3, 5, "cur"),
                            (4, 3, "yoy"), (4, 4, "ytd"), (4, 5, "mom")],
            }
            for r, c, key in header_map[sn]:
                ws.cell(r, c).value = labels[key]

        # Replace formula cells with computed values
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue

                # table 2 cross-refs: leave for second pass or resolve chain
                if "table " in v.lower():
                    continue  # second pass

                # multi price refs e.g. ='02'!$AD$738
                parts = re.findall(r"'([^']+)'!\$?([A-Z]+)\$?(\d+)", v)
                if not parts:
                    parts = re.findall(r"([A-Za-z0-9_ ]+)!\$?([A-Z]+)\$?(\d+)", v)
                if not parts:
                    continue

                # Single comparison ref
                if len(parts) == 1:
                    sheet, col, rnum = parts[0][0].strip(), parts[0][1], int(parts[0][2])
                    if col in COMP_COL:
                        val = _resolve_cell(sheet, col, rnum, lookup)
                        if val is not None:
                            cell.value = round(val, 1)
                            cell.number_format = "0.0"
                        else:
                            cell.value = None
                    else:
                        # price cell for table 11
                        # Map col letter to period via National month layout
                        # M=13=2023-01 ... offset from M
                        from openpyxl.utils import column_index_from_string

                        ci = column_index_from_string(col)
                        # month index 0 = 2023-01
                        mi = ci - 13
                        if 0 <= mi < len(result.months):
                            pkey = result.months[mi]["period"]
                            key = (sheet, rnum)
                            # rnum is price row
                            if key in prices and pkey in prices[key]:
                                pv = prices[key][pkey]
                                cell.value = round(pv, 1) if pv is not None else None
                                if pv is not None:
                                    cell.number_format = "#,##0.0"
                            else:
                                # try as aimag sheet price
                                cell.value = None
                        else:
                            cell.value = None
                else:
                    # multiple refs in one cell? unusual for tables; take first
                    sheet, col, rnum = parts[0][0].strip(), parts[0][1], int(parts[0][2])
                    if col in COMP_COL:
                        val = _resolve_cell(sheet, col, rnum, lookup)
                        cell.value = round(val, 1) if val is not None else None
                        if val is not None:
                            cell.number_format = "0.0"

    # Second pass: resolve remaining formulas (table→table, leftover specials)
    for sn in table_names:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                # ='table 6'!G6  or  =table 3!E6
                m = re.search(
                    r"table\s*(\d+)\s*'?!\$?([A-Z]+)\$?(\d+)", v, re.I
                )
                if m:
                    tsn = f"table {int(m.group(1))}"
                    col, rnum = m.group(2), int(m.group(3))
                    if tsn in wb.sheetnames:
                        ref_val = wb[tsn][f"{col}{rnum}"].value
                        if isinstance(ref_val, (int, float)):
                            cell.value = round(float(ref_val), 1)
                            cell.number_format = "0.0"
                    continue
                parts = re.findall(
                    r"'([^']+)'!\$?([A-Z]+)\$?(\d+)|([A-Za-z0-9_ ]+)!\$?([A-Z]+)\$?(\d+)",
                    v,
                )
                for p in parts:
                    if p[0]:
                        sheet, col, rnum = p[0].strip(), p[1], int(p[2])
                    else:
                        sheet, col, rnum = p[3].strip(), p[4], int(p[5])
                    if sheet.lower().startswith("table"):
                        continue
                    if col in COMP_COL:
                        val = _resolve_cell(sheet, col, rnum, lookup)
                        if val is not None:
                            cell.value = round(val, 1)
                            cell.number_format = "0.0"
                        break

    # Optional: keep only publication tables
    if tables_only:
        keep = set(table_names)
        for name in list(wb.sheetnames):
            if name not in keep:
                del wb[name]

    wb.save(output_path)
    return output_path


def publish_for_period(
    excel_path: str | Path,
    period: str | Period | None = None,
    template_path: str | Path | None = None,
    output_dir: str | Path = "output",
) -> Path:
    """Эхнээс нь: Excel унш → тооцоо → table 1–11."""
    from .loader import load_workbook_data
    from .engine import CPICalculator

    excel_path = Path(excel_path)
    data = load_workbook_data(excel_path)
    result = CPICalculator(data).calculate()

    if period is None:
        period = latest_period(result.months, result.national.overall)
    elif isinstance(period, str):
        period = parse_period(period)

    if template_path is None:
        template_path = Path(r"C:\Users\batsukh\Desktop\National_202607_2023.xlsx")
    template_path = Path(template_path)

    out = Path(output_dir) / f"National_{period.yyyymm}_2023.xlsx"
    return generate_publication(result, data, period, template_path, out, tables_only=True)
